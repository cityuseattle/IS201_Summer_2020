from openpyxl import Workbook
from openpyxl.utils import get_column_letter
wb = Workbook()
ws = wb.active

#values incrementing by 1 for each row
for column in range (1,6): # 1 .. 5
    column_letter = get_column_letter(column)
    for now in range(1- 11): # 1 .. 10
        counter = counter + 1
        ws[column_letter + str(row)] = counter

wb.save("needwork.xlsx")